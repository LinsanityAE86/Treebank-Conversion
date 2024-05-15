# # import openpyxl
# #
# # # 打开 Excel 文件
# # wb = openpyxl.Workbook()
# # sheet = wb.active
# #
# # # 打开要读取的文件
# # with open(r'D:\桌面\标签\UD-labels.txt', 'r',encoding='utf-8') as f:
# #     line = f.readline()
# #
# #     # 获取第一个 tab 的位置
# #     tab_index = line.index('\t')
# #
# #     # 将第一个 tab 前后的内容写入 Excel 表格的第一行和第二行
# #     sheet.cell(row=1, column=1).value = line[:tab_index]
# #     sheet.cell(row=2, column=1).value = line[tab_index + 1:]
# #
# # # 保存 Excel 文件
# # wb.save('output.xlsx')
# import openpyxl
#
# # 打开 Excel 文件
# wb = openpyxl.Workbook()
# sheet = wb.active
#
# # 打开要读取的文件
# with open(r'D:\桌面\标签\UD-labels.txt', 'r',encoding='utf-8') as f:
#     row_index = 1  # 初始化行索引
#
#     for line in f.readlines():
#         # 获取第一个 tab 的位置?
#         tab_index = line.index('\t')
#
#         # 将第一个 tab 前后的内容写入 Excel 表格的第一行和第二行
#         sheet.cell(row=row_index, column=1).value = line[:tab_index]
#         sheet.cell(row=row_index+1, column=1).value = line[tab_index + 1:]
#
#         # 更新行索引
#         row_index += 2
#
# # 保存 Excel 文件
# wb.save('output.xlsx')


import openpyxl

# 打开 Excel 文件
wb = openpyxl.Workbook()
sheet = wb.active

# 打开要读取的文件
with open(r'D:\桌面\标签\DP2020-labels.txt', 'r',encoding='utf-8') as f:
    row_index = 1  # 初始化行索引

    for line in f.readlines():
        # 获取第一个 tab 的位置
        tab_index = line.index('\t')

        # 将第一个 tab 前后的内容分别写入 Excel 表格的第一列和第二列
        sheet.cell(row=row_index, column=1).value = line[:tab_index]
        sheet.cell(row=row_index, column=2).value = line[tab_index + 1:]

        # 更新行索引
        row_index += 1

# 保存 Excel 文件
wb.save('output1.xlsx')
